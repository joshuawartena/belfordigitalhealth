"""
reporter.py
═══════════════════════════════════════════════════════════════════════════════
Digital Health Monitor — HTML / JSON / CSV / Excel Report Generator
═══════════════════════════════════════════════════════════════════════════════

Functions:
    generate_html_report  – Self-contained HTML report (Jinja2 inline template)
    generate_json_report  – Structured JSON export
    generate_batch_csv    – Multi-audit CSV summary
    generate_batch_excel  – Multi-audit XLSX with conditional formatting
"""

from __future__ import annotations

import base64
import csv
import json
import logging
import os
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from jinja2 import Template

from checkers import AuditResult, CategoryResult, CheckResult

logger = logging.getLogger('audit')


def _get_base64_asset(filename: str) -> str:
    """Read an asset file and return its base64 data URI."""
    try:
        # assets folder is located in same folder as reporter.py or adjacent
        asset_path = Path(__file__).parent / 'assets' / filename
        if not asset_path.exists():
            # Try just 'assets' from current working directory
            asset_path = Path('assets') / filename
        if asset_path.exists():
            ext = asset_path.suffix.lower().lstrip('.')
            if ext in ('jpg', 'jpeg'):
                mime = 'image/jpeg'
            elif ext == 'png':
                mime = 'image/png'
            elif ext == 'svg':
                mime = 'image/svg+xml'
            else:
                mime = 'application/octet-stream'
            
            data = asset_path.read_bytes()
            encoded = base64.b64encode(data).decode('utf-8')
            return f"data:{mime};base64,{encoded}"
    except Exception as e:
        logger.warning(f"Could not load asset {filename}: {e}")
    return ""

# ─── Color helpers ────────────────────────────────────────────────────────────

PRIORITY_COLORS = {
    'critical': '#dc2626',
    'high':     '#f97316',
    'medium':   '#f59e0b',
    'low':      '#3b82f6',
}


def _score_color(score: float) -> str:
    """Return a CSS color based on score thresholds."""
    if score >= 75:
        return '#22c55e'
    if score >= 60:
        return '#f59e0b'
    if score >= 40:
        return '#f97316'
    return '#ef4444'


def _score_bg(score: float) -> str:
    """Light background for score cells."""
    if score >= 75:
        return '#dcfce7'
    if score >= 60:
        return '#fef9c3'
    if score >= 40:
        return '#ffedd5'
    return '#fee2e2'


# ═══════════════════════════════════════════════════════════════════════════════
#  HTML Report
# ═══════════════════════════════════════════════════════════════════════════════

_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Digital Health Audit — {{ audit.name }}</title>
<style>
/* ── Reset & Base ─────────────────────────────────────────────────────── */
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body {
    font-family: 'Arial', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: #fdfdfc; color: #4a3e3d; line-height: 1.6; -webkit-font-smoothing: antialiased;
}
.container { max-width: 950px; margin: 0 auto; padding: 0 24px; }
a { color: #005DAA; text-decoration: none; }
a:hover { text-decoration: underline; }

/* ── Logo Bar ─────────────────────────────────────────────────────────── */
.logo-bar {
    background: #ffffff;
    padding: 14px 0;
    border-bottom: 3px solid #005DAA;
}
.logo-bar-inner {
    display: flex;
    justify-content: space-between;
    align-items: center;
}
.logo-bfg { height: 42px; width: auto; object-fit: contain; }
.logo-wms { height: 32px; width: auto; object-fit: contain; }
.logo-text { font-weight: bold; font-size: 15px; color: #8A7967; text-transform: uppercase; letter-spacing: 0.5px; }

/* ── Header (Centered Layout) ────────────────────────────────────────── */
.header {
    background: #005DAA; /* Flat solid Chem-Dry Blue */
    color: #ffffff;
    padding: 40px 0;
    text-align: center;
    position: relative;
}
.header-inner {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    gap: 16px;
}
.score-circle-container {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
}
.score-circle {
    width: 130px; height: 130px; border-radius: 50%; display: flex; flex-direction: column;
    align-items: center; justify-content: center; color: #ffffff; font-weight: bold;
    background: {{ grade_color }}; box-shadow: 0 0 0 6px rgba(255,255,255,0.2), 0 8px 20px rgba(0,0,0,0.2);
    position: relative;
}
.score-circle .grade  { font-size: 48px; line-height: 1; font-weight: bold; }
.score-circle .score  { font-size: 16px; opacity: 0.9; margin-top: -2px; }
.grade-label {
    text-align: center; margin-top: 10px; font-size: 15px; font-weight: bold;
    color: rgba(255, 255, 255, 0.95); text-shadow: 0 1px 2px rgba(0,0,0,0.15);
}
.header-text h1 { font-size: 30px; font-weight: bold; margin-bottom: 4px; color: #ffffff; letter-spacing: -0.5px; }
.header-text .domain-link { font-size: 17px; color: rgba(255,255,255,0.85); font-weight: 500; }
.header-text .domain-link:hover { color: #ffffff; }
.header-text .date   { font-size: 13px; color: rgba(255,255,255,0.65); margin-top: 4px; }

/* ── Cards Row ────────────────────────────────────────────────────────── */
.cards-row { display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin: -20px 0 32px; position: relative; z-index: 2; }
.card {
    background: #fff; border-radius: 8px; padding: 20px 24px;
    border: 1px solid #e5e2dc;
    box-shadow: 0 4px 12px rgba(0,0,0,0.03);
    transition: transform 0.2s, box-shadow 0.2s;
}
.card:hover { transform: translateY(-2px); box-shadow: 0 8px 20px rgba(0,0,0,0.06); }
.card-icon { font-size: 26px; margin-bottom: 6px; }
.card-name { font-size: 12px; font-weight: bold; color: #8A7967; text-transform: uppercase; letter-spacing: 0.5px; }
.card-score { font-size: 32px; font-weight: bold; margin: 4px 0 10px; }
.progress-track { background: #f0ede9; border-radius: 4px; height: 8px; overflow: hidden; }
.progress-fill { height: 100%; border-radius: 4px; }

/* ── Section ──────────────────────────────────────────────────────────── */
.section {
    background: #fff; border-radius: 8px; padding: 32px; margin-bottom: 24px;
    border: 1px solid #e5e2dc;
    box-shadow: 0 2px 8px rgba(0,0,0,0.02);
}
.section-title { font-size: 22px; font-weight: bold; margin-bottom: 20px; display: flex; align-items: center; gap: 10px; color: #005DAA; }
.section-title .icon { font-size: 24px; }

/* ── Recommendations ──────────────────────────────────────────────────── */
.rec-list { list-style: none; counter-reset: rec; }
.rec-item {
    display: flex; align-items: flex-start; gap: 14px; padding: 14px 0;
    border-bottom: 1px solid #f0ede9;
}
.rec-item:last-child { border-bottom: none; }
.rec-num {
    counter-increment: rec; width: 28px; height: 28px; border-radius: 50%;
    background: #f0ede9; display: flex; align-items: center; justify-content: center;
    font-size: 13px; font-weight: bold; color: #8A7967; flex-shrink: 0;
}
.rec-num::after { content: counter(rec); }
.priority-badge {
    display: inline-block; font-size: 10px; font-weight: bold; text-transform: uppercase;
    padding: 2px 8px; border-radius: 4px; color: #fff; letter-spacing: 0.3px; flex-shrink: 0;
    margin-top: 4px;
}
.rec-body .check-name { font-weight: bold; font-size: 14px; margin-bottom: 4px; color: #4a3e3d; }
.rec-body .rec-text   { font-size: 13px; color: #8A7967; line-height: 1.5; }

/* ── Detail Table ─────────────────────────────────────────────────────── */
.detail-table { width: 100%; border-collapse: separate; border-spacing: 0; font-size: 13px; margin-top: 10px; }
.detail-table thead th {
    background: #fdfdfc; text-align: left; padding: 12px 14px; font-weight: bold;
    color: #8A7967; text-transform: uppercase; font-size: 11px; letter-spacing: 0.4px;
    border-bottom: 2px solid #e5e2dc;
}
.detail-table tbody td { padding: 12px 14px; border-bottom: 1px solid #f0ede9; vertical-align: top; }
.detail-table tbody tr:last-child td { border-bottom: none; }

.row-pass { background: #f0fdf4; border-left: 4px solid #008752; }
.row-warn { background: #fffbeb; border-left: 4px solid #f59e0b; }
.row-fail { background: #fef2f2; border-left: 4px solid #dc2626; }
.row-error { background: #fef2f2; border-left: 4px solid #dc2626; }
.row-skip { background: #f8fafc; border-left: 4px solid #64748b; }

.status-icon { font-size: 16px; text-align: center; }
.points { font-weight: bold; white-space: nowrap; }
.value-cell { color: #4a3e3d; max-width: 250px; overflow: hidden; text-overflow: ellipsis; }
.detail-cell { color: #8A7967; line-height: 1.4; }

/* ── Category Header ──────────────────────────────────────────────────── */
.cat-header { display: flex; align-items: center; justify-content: space-between; margin-bottom: 20px; border-bottom: 2px solid #f0ede9; padding-bottom: 12px; }
.cat-header .cat-info { display: flex; align-items: center; gap: 10px; }
.cat-header .cat-icon { font-size: 28px; }
.cat-header .cat-name { font-size: 22px; font-weight: bold; color: #005DAA; }
.cat-score-badge {
    font-size: 16px; font-weight: bold; padding: 6px 14px; border-radius: 6px; color: #fff;
}

/* ── Footer ───────────────────────────────────────────────────────────── */
.footer {
    text-align: center; padding: 48px 0; color: #8A7967; font-size: 12px;
}
.footer .brand { font-weight: bold; color: #005DAA; }

/* ── Print ─────────────────────────────────────────────────────────────── */
@media print {
    body { background: #fff; }
    .card:hover { transform: none; box-shadow: none; }
    .section, .card { break-inside: avoid; box-shadow: none; border: 1px solid #e5e2dc; }
}
@media (max-width: 640px) {
    .header-inner { text-align: center; }
    .logo-bar-inner { flex-direction: column; gap: 12px; }
    .logo-bfg, .logo-wms { height: 30px; }
}
</style>
</head>
<body>

<!-- ═══ LOGO BAR ═══ -->
<div class="logo-bar">
    <div class="container logo-bar-inner">
        {% if belfor_logo %}
        <img src="{{ belfor_logo }}" class="logo-bfg" alt="BELFOR Franchise Group">
        {% else %}
        <span class="logo-text">BELFOR Franchise Group</span>
        {% endif %}
        
        {% if wms_logo %}
        <img src="{{ wms_logo }}" class="logo-wms" alt="WMS Logo">
        {% else %}
        <span class="logo-text">WMS</span>
        {% endif %}
    </div>
</div>

<!-- ═══ HEADER ═══ -->
<div class="header">
<div class="container header-inner">
    <div class="score-circle-container">
        <div class="score-circle">
            <span class="grade">{{ audit.grade }}</span>
            <span class="score">{{ audit.total_score | round(1) }}</span>
        </div>
        <div class="grade-label">{{ audit.grade_label }}</div>
    </div>
    
    <div class="header-text">
        <h1>{{ audit.name }}</h1>
        <div class="domain">
            <a href="{{ audit.url }}" target="_blank" class="domain-link">{{ audit.domain }}</a>
        </div>
        <div class="date">Digital Health Audit &bull; {{ audit.audited_at }}</div>
    </div>
</div>
</div>

<div class="container" style="margin-top: 24px;">

<!-- ═══ CATEGORY CARDS ═══ -->
<div class="cards-row">
{% for cat in categories %}
<div class="card">
    <div class="card-icon">{{ cat.icon }}</div>
    <div class="card-name">{{ cat.name }}</div>
    <div class="card-score" style="color:{{ score_color(cat.score) }}">{{ cat.score }}<span style="font-size:14px;color:#8A7967;font-weight:normal">/100</span></div>
    <div class="progress-track">
        <div class="progress-fill" style="width:{{ cat.score }}%;background:{{ score_color(cat.score) }}"></div>
    </div>
</div>
{% endfor %}
</div>

<!-- ═══ TOP RECOMMENDATIONS ═══ -->
{% if recommendations %}
<div class="section">
    <div class="section-title"><span class="icon">🎯</span> Top Recommendations</div>
    <ol class="rec-list">
    {% for rec in recommendations %}
        <li class="rec-item">
            <div class="rec-num"></div>
            <span class="priority-badge" style="background:{{ priority_color(rec.priority) }}">{{ rec.priority }}</span>
            <div class="rec-body">
                <div class="check-name">{{ rec.name }}</div>
                <div class="rec-text">{{ rec.recommendation }}</div>
            </div>
        </li>
    {% endfor %}
    </ol>
</div>
{% endif %}

<!-- ═══ DETAIL SECTIONS ═══ -->
{% for cat in categories %}
<div class="section">
    <div class="cat-header">
        <div class="cat-info">
            <span class="cat-icon">{{ cat.icon }}</span>
            <span class="cat-name">{{ cat.name }}</span>
        </div>
        <span class="cat-score-badge" style="background:{{ score_color(cat.score) }}">{{ cat.score }}/100</span>
    </div>
    <table class="detail-table">
        <thead>
            <tr>
                <th style="width:40px; text-align: center;">Status</th>
                <th>Check</th>
                <th>Value Found</th>
                <th>Detail</th>
                <th style="width:80px;text-align:right">Points</th>
            </tr>
        </thead>
        <tbody>
        {% for c in cat.checks %}
            <tr class="row-{{ c.status }}">
                <td class="status-icon">{{ c.icon }}</td>
                <td style="font-weight:bold; color: #4a3e3d;">{{ c.name }}</td>
                <td class="value-cell">{{ c.value or '—' }}</td>
                <td class="detail-cell">
                    <div>{{ c.detail or '—' }}</div>
                    {% if c.recommendation %}
                    <div style="margin-top: 6px; font-size: 12px; font-weight: 500; color: #4a3e3d; background: rgba(0, 0, 0, 0.02); padding: 6px; border-left: 2px solid #8A7967; border-radius: 2px;">
                        <strong>Recommendation:</strong> {{ c.recommendation }}
                    </div>
                    {% endif %}
                </td>
                <td class="points" style="text-align:right;color:{{ c.status_color }}">{{ c.points_earned }}/{{ c.points_possible }}</td>
            </tr>
        {% endfor %}
        </tbody>
    </table>
</div>
{% endfor %}

<!-- ═══ FOOTER ═══ -->
<div class="footer">
    <div>Generated by <span class="brand">Digital Health Monitor</span> · {{ timestamp }}</div>
    <div style="margin-top:4px">BFG WMS — Franchise SEO Audit</div>
</div>

</div><!-- .container -->
</body>
</html>"""


def generate_html_report(audit: AuditResult, output_path: str) -> str:
    """
    Generate a fully self-contained HTML audit report.

    Uses Jinja2 with an inline template.  The output file embeds all CSS and
    can be emailed or opened directly in a browser.

    Args:
        audit: Completed AuditResult with scored categories.
        output_path: Filesystem path for the output .html file.

    Returns:
        The *output_path* that was written.
    """
    try:
        template = Template(_HTML_TEMPLATE)
        categories = audit.categories
        recommendations = audit.top_recommendations(10)

        html = template.render(
            audit=audit,
            categories=categories,
            recommendations=recommendations,
            grade_color=audit.grade_color or '#3b82f6',
            score_color=_score_color,
            priority_color=lambda p: PRIORITY_COLORS.get(p, '#64748b'),
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M'),
            belfor_logo=_get_base64_asset('BELFOR_Franchise_Group_.jpg'),
            wms_logo=_get_base64_asset('WMS-Logo-Transparent-cropped.png'),
        )

        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        Path(output_path).write_text(html, encoding='utf-8')
        logger.info('HTML report written → %s', output_path)
        return output_path

    except Exception:
        logger.exception('Failed to generate HTML report')
        raise


# ═══════════════════════════════════════════════════════════════════════════════
#  JSON Report
# ═══════════════════════════════════════════════════════════════════════════════

def _check_to_dict(check: CheckResult) -> Dict[str, Any]:
    """Serialize a CheckResult including computed properties."""
    d = asdict(check)
    d['icon'] = check.icon
    d['status_color'] = check.status_color
    return d


def _category_to_dict(cat: CategoryResult) -> Dict[str, Any]:
    """Serialize a CategoryResult including computed properties."""
    return {
        'name': cat.name,
        'key': cat.key,
        'icon': cat.icon,
        'score': cat.score,
        'earned': cat.earned,
        'possible': cat.possible,
        'pass_count': cat.pass_count,
        'warn_count': cat.warn_count,
        'fail_count': cat.fail_count,
        'checks': [_check_to_dict(c) for c in cat.checks],
    }


def _audit_to_dict(audit: AuditResult) -> Dict[str, Any]:
    """Serialize a full AuditResult to a plain dict."""
    return {
        'name': audit.name,
        'domain': audit.domain,
        'url': audit.url,
        'city': audit.city,
        'state': audit.state,
        'phone': audit.phone,
        'address': audit.address,
        'zip_code': audit.zip_code,
        'owner': audit.owner,
        'place_id': audit.place_id,
        'gbp_url': audit.gbp_url,
        'business_name': audit.business_name,
        'audited_at': audit.audited_at,
        'http_status': audit.http_status,
        'error': audit.error,
        'total_score': audit.total_score,
        'grade': audit.grade,
        'grade_label': audit.grade_label,
        'grade_color': audit.grade_color,
        'total_pass': audit.total_pass,
        'total_warn': audit.total_warn,
        'total_fail': audit.total_fail,
        'categories': {
            cat.key: _category_to_dict(cat) for cat in audit.categories
        },
        'top_recommendations': [
            _check_to_dict(c) for c in audit.top_recommendations(10)
        ],
    }


def generate_json_report(audit: AuditResult, output_path: str) -> str:
    """
    Export the audit as a structured JSON file.

    Includes computed properties (scores, counts, icons) that are not part
    of the raw dataclass fields.

    Args:
        audit: Completed AuditResult.
        output_path: Filesystem path for the output .json file.

    Returns:
        The *output_path* that was written.
    """
    try:
        data = _audit_to_dict(audit)
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        Path(output_path).write_text(
            json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8',
        )
        logger.info('JSON report written → %s', output_path)
        return output_path

    except Exception:
        logger.exception('Failed to generate JSON report')
        raise


# ═══════════════════════════════════════════════════════════════════════════════
#  Batch CSV
# ═══════════════════════════════════════════════════════════════════════════════

_CSV_COLUMNS = [
    'name', 'domain', 'city', 'state',
    'total_score', 'grade',
    'website_seo_score', 'pagespeed_score', 'gbp_score',
    'total_pass', 'total_warn', 'total_fail',
]


def _audit_row(audit: AuditResult) -> Dict[str, Any]:
    """Extract a flat row dict from an AuditResult."""
    return {
        'name':              audit.name,
        'domain':            audit.domain,
        'city':              audit.city,
        'state':             audit.state,
        'total_score':       round(audit.total_score, 1),
        'grade':             audit.grade,
        'website_seo_score': audit.website_seo.score if audit.website_seo else '',
        'pagespeed_score':   audit.pagespeed.score if audit.pagespeed else '',
        'gbp_score':         audit.gbp.score if audit.gbp else '',
        'total_pass':        audit.total_pass,
        'total_warn':        audit.total_warn,
        'total_fail':        audit.total_fail,
    }


def generate_batch_csv(audits: list, output_path: str) -> str:
    """
    Generate a CSV summary with one row per audit.

    Columns: name, domain, city, state, total_score, grade,
             website_seo_score, pagespeed_score, gbp_score,
             total_pass, total_warn, total_fail.

    Args:
        audits: List of AuditResult instances.
        output_path: Filesystem path for the output .csv file.

    Returns:
        The *output_path* that was written.
    """
    try:
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        with open(output_path, 'w', newline='', encoding='utf-8') as fh:
            writer = csv.DictWriter(fh, fieldnames=_CSV_COLUMNS)
            writer.writeheader()
            for audit in audits:
                writer.writerow(_audit_row(audit))
        logger.info('Batch CSV written → %s  (%d audits)', output_path, len(audits))
        return output_path

    except Exception:
        logger.exception('Failed to generate batch CSV')
        raise


# ═══════════════════════════════════════════════════════════════════════════════
#  Batch Excel
# ═══════════════════════════════════════════════════════════════════════════════

def generate_batch_excel(audits: list, output_path: str) -> str:
    """
    Generate an Excel workbook summary with one row per audit.

    Features:
        • Bold header row with blue background (#3b82f6) and white font.
        • Score columns conditionally coloured (green ≥75, yellow ≥60,
          orange ≥40, red <40).
        • Auto-sized columns for readability.

    Args:
        audits: List of AuditResult instances.
        output_path: Filesystem path for the output .xlsx file.

    Returns:
        The *output_path* that was written.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Audit Summary'

        # ── Header style ──
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')

        # Write headers
        for col_idx, col_name in enumerate(_CSV_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # ── Score fill colours ──
        fill_green  = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
        fill_yellow = PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid')
        fill_orange = PatternFill(start_color='FFEDD5', end_color='FFEDD5', fill_type='solid')
        fill_red    = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

        score_cols = {'total_score', 'website_seo_score', 'pagespeed_score', 'gbp_score'}

        # Write data rows
        for row_idx, audit in enumerate(audits, start=2):
            row = _audit_row(audit)
            for col_idx, col_name in enumerate(_CSV_COLUMNS, start=1):
                val = row[col_name]
                cell = ws.cell(row=row_idx, column=col_idx, value=val)

                # Conditional score colouring
                if col_name in score_cols and isinstance(val, (int, float)):
                    if val >= 75:
                        cell.fill = fill_green
                    elif val >= 60:
                        cell.fill = fill_yellow
                    elif val >= 40:
                        cell.fill = fill_orange
                    else:
                        cell.fill = fill_red

        # ── Auto-size columns ──
        for col_idx in range(1, len(_CSV_COLUMNS) + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        max_len = max(max_len, len(str(cell.value or '')))
                    except Exception:
                        pass
            ws.column_dimensions[col_letter].width = max_len + 4

        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        wb.save(output_path)
        logger.info('Batch Excel written → %s  (%d audits)', output_path, len(audits))
        return output_path

    except Exception:
        logger.exception('Failed to generate batch Excel report')
        raise
